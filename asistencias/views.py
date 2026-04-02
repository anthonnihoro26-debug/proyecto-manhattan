import json
import logging
import math
import re
from datetime import datetime, time, timedelta
from io import BytesIO, StringIO

from PIL import Image as PILImage
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login as auth_login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.staticfiles import finders
from django.core.management import call_command
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Min, Q
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.csrf import csrf_exempt, csrf_protect, ensure_csrf_cookie
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from .models import Asistencia, JustificacionAsistencia, Profesor, DiaEspecial

logger = logging.getLogger(__name__)

try:
    import cloudinary.uploader
    CLOUDINARY_AVAILABLE = True
except Exception:
    CLOUDINARY_AVAILABLE = False


# =========================================================
# HELPERS DE ROLES
# =========================================================
def _in_group(group_name: str):
    def check(user):
        return user.is_authenticated and (
            user.is_superuser or user.groups.filter(name=group_name).exists()
        )
    return check


def _in_any_group(*group_names: str):
    def check(user):
        return user.is_authenticated and (
            user.is_superuser or user.groups.filter(name__in=list(group_names)).exists()
        )
    return check


def _is_private_owner(user):
    return (
        user.is_authenticated
        and user.is_superuser
        and user.username.lower() == "anthonny"
    )


GROUP_DESTINATIONS = {
    "SCANNER": "scan_page",
    "HISTORIAL": "historial_asistencias",
    "JUSTIFICACIONES": "panel_justificaciones",
}


def get_user_allowed_groups(user):
    if not user.is_authenticated:
        return []

    if user.username.lower() == "anthonny" and user.is_superuser:
        return []

    if user.is_superuser:
        return list(GROUP_DESTINATIONS.keys())

    user_groups = list(user.groups.values_list("name", flat=True))
    return [g for g in user_groups if g in GROUP_DESTINATIONS]


def _get_client_ip(request):
    xff = request.META.get("HTTP_X_FORWARDED_FOR")
    if xff:
        return xff.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "")


def _extract_dni(raw: str) -> str:
    raw = (raw or "").strip()

    match = re.search(r"(\d{8})", raw)
    if match:
        return match.group(1)

    digits = re.sub(r"\D+", "", raw)
    if len(digits) == 7:
        return digits.zfill(8)
    if len(digits) >= 8:
        return digits[-8:]
    return ""


def _read_code_from_request(request) -> str:
    ctype = (request.content_type or "").lower()

    if "application/json" in ctype:
        body = (request.body or b"").decode("utf-8").strip()
        if not body:
            return ""
        data = json.loads(body)
        return (data.get("code") or data.get("dni") or "").strip()

    return (request.POST.get("code") or request.POST.get("dni") or "").strip()


def _safe_file_url(file_field):
    try:
        if file_field:
            return file_field.url
    except Exception:
        return None
    return None


# =========================================================
# HELPERS DÍAS ESPECIALES
# =========================================================
def _dias_especiales_dict(fecha_inicio, fecha_fin):
    qs = (
        DiaEspecial.objects
        .filter(activo=True, fecha__range=(fecha_inicio, fecha_fin))
        .only("fecha", "tipo", "descripcion", "activo")
        .order_by("-fecha")
    )

    out = {}
    for d in qs:
        try:
            tipo_display = d.get_tipo_display()
        except Exception:
            tipo_display = (d.tipo or "").replace("_", " ").title()

        descripcion = (d.descripcion or "").strip()
        label = tipo_display
        if descripcion:
            label = f"{tipo_display} - {descripcion}"

        out[d.fecha] = {
            "tipo": d.tipo,
            "tipo_display": tipo_display,
            "descripcion": descripcion,
            "label": label,
        }
    return out


def _es_dia_especial(fecha):
    return (
        DiaEspecial.objects
        .filter(fecha=fecha, activo=True)
        .exists()
    )


def _obtener_dia_especial(fecha):
    return (
        DiaEspecial.objects
        .filter(fecha=fecha, activo=True)
        .first()
    )


def _tipo_display_dia_especial(dia_especial):
    try:
        return dia_especial.get_tipo_display()
    except Exception:
        return (dia_especial.tipo or "").replace("_", " ").title()


# =========================================================
# HELPERS GEOLOGIN
# =========================================================
JORGE_GEOFENCE_USERNAME = "jorge"
JORGE_GEOFENCE_LAT = -12.0360672
JORGE_GEOFENCE_LNG = -77.0033333
JORGE_GEOFENCE_RADIUS_M = 30.0


def _to_float_maybe(value):
    try:
        if value is None:
            return None
        s = str(value).strip()
        if not s:
            return None
        s = s.replace(",", ".")
        return float(s)
    except (ValueError, TypeError):
        return None


def _haversine_m(lat1, lon1, lat2, lon2):
    radius = 6371000.0
    p1 = math.radians(float(lat1))
    p2 = math.radians(float(lat2))
    dp = math.radians(float(lat2) - float(lat1))
    dl = math.radians(float(lon2) - float(lon1))

    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return radius * c


# =========================================================
# LOGIN
# =========================================================
def login_view_geocerca(request):
    if request.user.is_authenticated:
        return redirect("post_login")

    form = AuthenticationForm(request=request, data=request.POST or None)

    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        username_lc = username.lower()

        geo_status = (request.POST.get("geo_status") or "").strip().lower()
        geo_lat = _to_float_maybe(request.POST.get("geo_lat"))
        geo_lng = _to_float_maybe(request.POST.get("geo_lng"))
        geo_acc = _to_float_maybe(request.POST.get("geo_acc"))

        if username_lc == JORGE_GEOFENCE_USERNAME:
            if geo_status != "ok" or geo_lat is None or geo_lng is None:
                messages.error(
                    request,
                    "Jorge: debes permitir la ubicación para iniciar sesión."
                )
                return render(request, "login.html", {"form": form})

            distancia_m = _haversine_m(
                geo_lat,
                geo_lng,
                JORGE_GEOFENCE_LAT,
                JORGE_GEOFENCE_LNG,
            )

            if distancia_m > JORGE_GEOFENCE_RADIUS_M:
                msg = (
                    f"Jorge: fuera del área permitida. "
                    f"Distancia detectada: {distancia_m:.1f} m "
                    f"(máximo {JORGE_GEOFENCE_RADIUS_M:.0f} m)."
                )
                if geo_acc is not None:
                    msg += f" Precisión reportada: ±{geo_acc:.0f} m."
                messages.error(request, msg)
                return render(request, "login.html", {"form": form})

        if form.is_valid():
            user = form.get_user()
            auth_login(request, user)
            return redirect(request.POST.get("next") or "post_login")

    return render(request, "login.html", {"form": form})


# =========================================================
# HELPERS HISTORIAL POR DÍA
# =========================================================
def _aware_midnight(d):
    dt = datetime.combine(d, time(0, 0, 0))
    tz = timezone.get_current_timezone()
    return timezone.make_aware(dt, tz)


def _aware_end_of_day(d):
    dt = datetime.combine(d, time(23, 59, 59))
    tz = timezone.get_current_timezone()
    return timezone.make_aware(dt, tz)


def _build_historial_rows_por_dia(fecha, q="", condicion=""):
    profesores_qs = (
        Profesor.objects.only("id", "dni", "codigo", "apellidos", "nombres", "condicion")
        .order_by("apellidos", "nombres")
    )

    if q:
        profesores_qs = profesores_qs.filter(
            Q(dni__icontains=q)
            | Q(codigo__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(nombres__icontains=q)
        )

    if condicion in ("N", "C"):
        profesores_qs = profesores_qs.filter(condicion__iexact=condicion)

    profesores = list(profesores_qs)
    profesor_ids = [p.id for p in profesores]

    dia_especial = _obtener_dia_especial(fecha)

    asistencias = (
        Asistencia.objects
        .filter(profesor_id__in=profesor_ids, fecha=fecha, tipo="E")
        .select_related("profesor")
        .order_by("fecha_hora")
    )
    asistencia_map = {}
    for a in asistencias:
        if a.profesor_id not in asistencia_map:
            asistencia_map[a.profesor_id] = a

    justificaciones = (
        JustificacionAsistencia.objects
        .filter(profesor_id__in=profesor_ids, fecha=fecha)
        .select_related("profesor")
    )
    just_map = {j.profesor_id: j for j in justificaciones}

    asist_j = (
        Asistencia.objects
        .filter(profesor_id__in=profesor_ids, fecha=fecha, tipo="J")
        .select_related("profesor")
        .order_by("fecha_hora")
    )
    asist_j_map = {}
    for j in asist_j:
        if j.profesor_id not in asist_j_map:
            asist_j_map[j.profesor_id] = j

    rows = []
    c_asistio = 0
    c_just = 0
    c_falto = 0
    c_especial = 0

    for profesor in profesores:
        a = asistencia_map.get(profesor.id)
        j = just_map.get(profesor.id)
        aj = asist_j_map.get(profesor.id)

        if dia_especial:
            tipo_display = _tipo_display_dia_especial(dia_especial)
            descripcion = (dia_especial.descripcion or "").strip()

            row = {
                "profesor": profesor,
                "estado_key": "DIA_ESPECIAL",
                "estado": tipo_display.upper(),
                "detalle": descripcion or "Día especial institucional",
                "fecha_hora": _aware_end_of_day(fecha),
                "puede_justificar": False,
                "justificacion_existente": False,
                "es_dia_especial": True,
            }
            c_especial += 1

        elif a:
            row = {
                "profesor": profesor,
                "estado_key": "ASISTIO",
                "estado": "ASISTIÓ",
                "detalle": "Asistencia registrada",
                "fecha_hora": a.fecha_hora,
                "puede_justificar": False,
                "justificacion_existente": bool(j or aj),
                "es_dia_especial": False,
            }
            c_asistio += 1

        elif j or aj:
            motivo_label = "Justificación"
            detalle = ""

            if j:
                try:
                    motivo_label = j.get_tipo_display()
                except Exception:
                    motivo_label = (j.tipo or "").strip() or "Justificación"
                detalle = (j.detalle or "").strip()
                fecha_hora = _aware_midnight(j.fecha)
            else:
                try:
                    motivo_label = aj.get_motivo_display()
                except Exception:
                    motivo_label = (getattr(aj, "motivo", "") or "").strip() or "Justificación"
                detalle = (getattr(aj, "detalle", "") or "").strip()
                fecha_hora = aj.fecha_hora

            row = {
                "profesor": profesor,
                "estado_key": "JUSTIFICADO",
                "estado": f"JUSTIFICADO ({motivo_label})",
                "detalle": detalle or "Inasistencia justificada",
                "fecha_hora": fecha_hora,
                "puede_justificar": False,
                "justificacion_existente": True,
                "es_dia_especial": False,
            }
            c_just += 1

        else:
            row = {
                "profesor": profesor,
                "estado_key": "FALTO",
                "estado": "FALTÓ",
                "detalle": "Sin asistencia ni justificación",
                "fecha_hora": _aware_end_of_day(fecha),
                "puede_justificar": True,
                "justificacion_existente": False,
                "es_dia_especial": False,
            }
            c_falto += 1

        rows.append(row)

    return {
        "rows": rows,
        "dia_especial": dia_especial,
        "resumen": {
            "asistio": c_asistio,
            "justificado": c_just,
            "falto": c_falto,
            "especial": c_especial,
            "total": len(rows),
        },
    }


# =========================================================
# POST LOGIN
# =========================================================
@login_required
def post_login_redirect(request):
    user = request.user

    if user.username.lower() == "anthonny" and user.is_superuser:
        return redirect("estadisticas_privadas")

    allowed_groups = get_user_allowed_groups(user)

    if not allowed_groups:
        logout(request)
        messages.error(request, "Tu usuario no tiene módulos asignados.")
        return redirect("login")

    if len(allowed_groups) == 1:
        selected_group = allowed_groups[0]
        request.session["selected_group"] = selected_group
        return redirect(GROUP_DESTINATIONS[selected_group])

    return redirect("seleccionar_grupo")


# =========================================================
# SELECTOR DE MÓDULO
# =========================================================
@login_required
def seleccionar_grupo(request):
    allowed_groups = get_user_allowed_groups(request.user)

    if not allowed_groups:
        logout(request)
        messages.error(request, "Tu usuario no tiene módulos asignados.")
        return redirect("login")

    if len(allowed_groups) == 1:
        group_name = allowed_groups[0]
        request.session["selected_group"] = group_name
        return redirect(GROUP_DESTINATIONS[group_name])

    if request.method == "POST":
        grupo_elegido = (request.POST.get("grupo") or "").strip()

        if grupo_elegido not in allowed_groups:
            messages.error(request, "Grupo no válido.")
            return redirect("seleccionar_grupo")

        request.session["selected_group"] = grupo_elegido
        messages.success(request, f"Ingresaste al módulo: {grupo_elegido}")
        return redirect(GROUP_DESTINATIONS[grupo_elegido])

    return render(
        request,
        "asistencias/seleccionar_grupo.html",
        {"allowed_groups": allowed_groups},
    )


# =========================================================
# HISTORIAL POR DÍA
# =========================================================
@user_passes_test(_in_any_group("HISTORIAL", "JUSTIFICACIONES"), login_url="login")
def historial_asistencias(request):
    q = (request.GET.get("q") or "").strip()
    fecha_str = (request.GET.get("fecha") or "").strip()
    condicion = (request.GET.get("condicion") or "").strip().upper()
    ps = (request.GET.get("ps") or "25").strip()

    if ps not in ("25", "50", "100"):
        ps = "25"
    ps = int(ps)

    fecha = parse_date(fecha_str) if fecha_str else timezone.localdate()
    if not fecha:
        fecha = timezone.localdate()

    origen = (request.GET.get("from") or "").strip().lower()

    if origen == "justificaciones":
        request.session["historial_origen"] = "justificaciones"
    elif origen == "historial":
        request.session["historial_origen"] = "historial"

    historial_origen = request.session.get("historial_origen", "historial")

    fecha_just = request.session.get("just_fecha")
    if not fecha_just:
        fecha_just = (timezone.localdate() - timedelta(days=1)).strftime("%Y-%m-%d")

    puede_volver_just = (
        request.user.is_authenticated
        and (
            request.user.is_superuser
            or request.user.groups.filter(name="JUSTIFICACIONES").exists()
        )
        and historial_origen == "justificaciones"
    )

    if historial_origen == "justificaciones":
        url_registro_manual = f"{reverse('registro_manual')}?from=justificaciones"
    else:
        url_registro_manual = f"{reverse('registro_manual')}?from=historial"

    url_volver_just = f"{reverse('panel_justificaciones')}?fecha={fecha_just}"

    data = _build_historial_rows_por_dia(
        fecha=fecha,
        q=q,
        condicion=condicion,
    )

    rows = data["rows"]
    dia_especial = data["dia_especial"]
    resumen = data["resumen"]

    paginator = Paginator(rows, ps)
    page_number = request.GET.get("page", "1")
    page_obj = paginator.get_page(page_number)
    items = list(page_obj.object_list)

    can_justify_from_historial = (
        request.user.is_superuser
        or request.user.groups.filter(name="HISTORIAL").exists()
        or request.user.groups.filter(name="JUSTIFICACIONES").exists()
    )

    return render(
        request,
        "asistencias/historial.html",
        {
            "items": items,
            "page_obj": page_obj,
            "paginator": paginator,
            "q": q,
            "fecha": fecha,
            "condicion": condicion,
            "ps": ps,
            "total_registros": resumen["total"],
            "total_asist": resumen["asistio"],
            "total_just": resumen["justificado"],
            "total_falto": resumen["falto"],
            "total_especiales": resumen["especial"],
            "docentes_unicos": resumen["total"],
            "registros_n": len([x for x in rows if (x["profesor"].condicion or "").upper() == "N"]),
            "registros_c": len([x for x in rows if (x["profesor"].condicion or "").upper() == "C"]),
            "puede_volver_just": puede_volver_just,
            "fecha_just": fecha_just,
            "url_registro_manual": url_registro_manual,
            "url_volver_just": url_volver_just,
            "dia_especial": dia_especial,
            "can_justify_from_historial": can_justify_from_historial,
        },
    )


# =========================================================
# JUSTIFICAR FALTA DESDE HISTORIAL
# =========================================================
@require_POST
@user_passes_test(_in_any_group("HISTORIAL", "JUSTIFICACIONES"), login_url="login")
def justificar_falta_historial(request):
    profesor_id = (request.POST.get("profesor_id") or "").strip()
    fecha_str = (request.POST.get("fecha") or "").strip()
    tipo = (request.POST.get("tipo") or "DM").strip().upper()
    detalle = (request.POST.get("detalle") or "").strip()

    q = (request.POST.get("q") or "").strip()
    condicion = (request.POST.get("condicion") or "").strip()
    ps = (request.POST.get("ps") or "25").strip()
    page = (request.POST.get("page") or "1").strip()

    redirect_url = (
        f"{reverse('historial_asistencias')}?"
        f"q={q}&fecha={fecha_str}&condicion={condicion}&ps={ps}&page={page}"
    )

    fecha = parse_date(fecha_str)
    if not fecha:
        messages.error(request, "Fecha inválida.")
        return redirect(redirect_url)

    if _es_dia_especial(fecha):
        messages.warning(request, "Ese día está marcado como día especial. No se requiere justificación.")
        return redirect(redirect_url)

    try:
        profesor = Profesor.objects.get(id=profesor_id)
    except Profesor.DoesNotExist:
        messages.error(request, "Profesor no encontrado.")
        return redirect(redirect_url)

    tipo_ok = tipo if tipo in ("DM", "C", "P", "O") else "DM"
    ip = _get_client_ip(request)
    ua = (request.META.get("HTTP_USER_AGENT") or "")[:255]

    if Asistencia.objects.filter(profesor=profesor, fecha=fecha, tipo="E").exists():
        messages.warning(request, "Ese docente ya tiene asistencia registrada en esa fecha.")
        return redirect(redirect_url)

    if JustificacionAsistencia.objects.filter(profesor=profesor, fecha=fecha).exists():
        messages.warning(request, "Ese docente ya tiene justificación registrada en esa fecha.")
        return redirect(redirect_url)

    try:
        with transaction.atomic():
            JustificacionAsistencia.objects.create(
                profesor=profesor,
                fecha=fecha,
                tipo=tipo_ok,
                detalle=detalle,
                creado_por=request.user,
                actualizado_por=request.user,
            )

            Asistencia.objects.update_or_create(
                profesor=profesor,
                fecha=fecha,
                tipo="J",
                defaults={
                    "fecha_hora": timezone.now(),
                    "motivo": tipo_ok,
                    "detalle": detalle,
                    "registrado_por": request.user,
                    "ip": ip,
                    "user_agent": ua,
                },
            )

        messages.success(request, f"✅ Falta justificada correctamente para {profesor.apellidos}, {profesor.nombres}.")
        return redirect(redirect_url)

    except IntegrityError:
        messages.warning(request, "Ya existía una justificación para ese docente en esa fecha.")
        return redirect(redirect_url)

    except Exception as e:
        messages.error(request, f"Error guardando la justificación: {type(e).__name__} - {str(e)[:220]}")
        return redirect(redirect_url)


# =========================================================
# EXCEL REPORTE GENERAL
# =========================================================
@user_passes_test(_in_any_group("HISTORIAL", "JUSTIFICACIONES"), login_url="login")
def exportar_reporte_excel(request):
    q = (request.GET.get("q") or "").strip()
    condicion = (request.GET.get("condicion") or "").strip().upper()
    fecha_str = (request.GET.get("fecha") or "").strip()

    hoy = timezone.localdate()
    fecha = parse_date(fecha_str) if fecha_str else hoy
    if not fecha:
        fecha = hoy

    desde = fecha
    hasta = fecha

    dias_especiales = _dias_especiales_dict(desde, hasta)

    profesores_qs = Profesor.objects.all().order_by("apellidos", "nombres")

    if q:
        profesores_qs = profesores_qs.filter(
            Q(dni__icontains=q)
            | Q(codigo__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(nombres__icontains=q)
        )

    if condicion in ("N", "C"):
        profesores_qs = profesores_qs.filter(condicion__iexact=condicion)

    profesores = list(profesores_qs)
    prof_ids = [p.id for p in profesores]

    entradas = (
        Asistencia.objects.filter(
            profesor_id__in=prof_ids,
            fecha=fecha,
            tipo="E",
        )
        .values("profesor_id", "fecha")
        .annotate(primera_hora=Min("fecha_hora"))
    )
    entrada_map = {(x["profesor_id"], x["fecha"]): x["primera_hora"] for x in entradas}

    justificados = (
        JustificacionAsistencia.objects.filter(
            profesor_id__in=prof_ids,
            fecha=fecha,
        )
        .values("profesor_id", "fecha", "tipo", "detalle")
    )

    motivos_label = {
        "DM": "Descanso médico",
        "C": "Comisión / Encargo",
        "P": "Permiso",
        "O": "Otro",
    }

    just_map = {}
    for j in justificados:
        key = (j["profesor_id"], j["fecha"])
        t = (j.get("tipo") or "").strip()
        det = (j.get("detalle") or "").strip()
        label = motivos_label.get(t, t or "Justificación")
        just_map[key] = f"JUSTIFICADO ({label})" + (f" - {det}" if det else "")

    asist_j = (
        Asistencia.objects.filter(
            profesor_id__in=prof_ids,
            fecha=fecha,
            tipo="J",
        )
        .values("profesor_id", "fecha", "motivo", "detalle")
    )

    asist_j_map = {}
    for a in asist_j:
        key = (a["profesor_id"], a["fecha"])
        mot = (a.get("motivo") or "").strip()
        det = (a.get("detalle") or "").strip()
        label = motivos_label.get(mot, mot or "Justificación")
        asist_j_map[key] = f"JUSTIFICADO ({label})" + (f" - {det}" if det else "")

    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Reporte Diario"

    navy = "7F1D1D"
    red = "B91C1C"
    red_soft = "FEE2E2"
    green_soft = "DCFCE7"
    blue_soft = "DBEAFE"
    amber_soft = "FEF3C7"
    gray_bg = "F8FAFC"
    text_dark = "111827"
    white = "FFFFFF"

    thin = Side(style="thin", color="CBD5E1")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_columns = 7
    last_col_letter = get_column_letter(total_columns)

    for r in (1, 2):
        for c in range(1, total_columns + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=gray_bg)

    logo_path = finders.find("asistencias/img/uni_logo.png")
    if logo_path:
        with PILImage.open(logo_path) as im:
            ow, oh = im.size

        target_h = 95
        target_w = int(ow * (target_h / oh))
        img = XLImage(logo_path)
        img.height = target_h
        img.width = target_w
        img.anchor = "A1"
        ws.add_image(img)

        ws.row_dimensions[1].height = 44
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 10
        ws.column_dimensions["A"].width = 18

    titulo = f"REPORTE DIARIO DE ASISTENCIAS — {fecha.strftime('%d/%m/%Y')}"
    ws["B1"] = titulo
    ws.merge_cells(f"B1:{last_col_letter}1")
    ws["B1"].font = Font(bold=True, size=16, color=navy)
    ws["B1"].alignment = Alignment(vertical="center")

    filtros_txt = []
    if q:
        filtros_txt.append(f"Búsqueda: {q}")
    if condicion:
        filtros_txt.append(f"Condición: {condicion.upper()}")
    filtros_txt.append(f"Fecha: {fecha.strftime('%Y-%m-%d')}")
    filtros_txt.append(f"Docentes: {len(profesores)}")

    ws["B2"] = " | ".join(filtros_txt)
    ws.merge_cells(f"B2:{last_col_letter}2")
    ws["B2"].font = Font(size=11, color="334155")
    ws["B2"].alignment = Alignment(vertical="center")

    ws.append([])

    headers = ["DNI", "Código", "Docente", "Condición", "Estado", "Detalle", "Hora"]
    ws.append([str(h) for h in headers])
    header_row = ws.max_row

    header_fill = PatternFill("solid", fgColor=red)
    header_font = Font(bold=True, color=white)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all

    dia_especial = _obtener_dia_especial(fecha)

    for p in profesores:
        key = (p.id, fecha)
        dt = entrada_map.get(key)

        if dia_especial:
            estado = _tipo_display_dia_especial(dia_especial).upper()
            detalle = (dia_especial.descripcion or "").strip() or "Día especial institucional"
            hora = "-"
        elif dt:
            dt_local = timezone.localtime(dt)
            estado = "ASISTIÓ"
            detalle = "Asistencia registrada"
            hora = dt_local.strftime("%H:%M")
        else:
            jtxt = just_map.get(key) or asist_j_map.get(key)
            if jtxt:
                estado = "JUSTIFICADO"
                detalle = jtxt
                hora = "-"
            else:
                estado = "FALTÓ"
                detalle = "Sin asistencia ni justificación"
                hora = "-"

        docente = f"{(p.apellidos or '').strip()}, {(p.nombres or '').strip()}".strip().strip(",")

        row = [
            str(p.dni),
            str(p.codigo or ""),
            docente,
            str((p.condicion or "").upper()),
            estado,
            detalle,
            hora,
        ]
        ws.append(row)

        current_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border_all
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        estado_cell = ws.cell(row=current_row, column=5)
        if estado == "ASISTIÓ":
            estado_cell.fill = PatternFill("solid", fgColor=green_soft)
        elif estado == "JUSTIFICADO":
            estado_cell.fill = PatternFill("solid", fgColor=blue_soft)
        elif estado == "FALTÓ":
            estado_cell.fill = PatternFill("solid", fgColor=red_soft)
        else:
            estado_cell.fill = PatternFill("solid", fgColor=amber_soft)

    ws.freeze_panes = "A5"

    widths = [14, 14, 40, 12, 18, 45, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    filename = f"reporte_diario_asistencias_{fecha.strftime('%Y-%m-%d')}.xlsx"
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# =========================================================
# SCAN
# =========================================================
@ensure_csrf_cookie
@user_passes_test(_in_group("SCANNER"), login_url="login")
def scan_page(request):
    return render(request, "asistencias/scan.html")


@csrf_protect
@require_POST
@user_passes_test(_in_group("SCANNER"), login_url="login")
def api_scan_asistencia(request):
    try:
        raw = _read_code_from_request(request)
    except UnicodeDecodeError:
        return JsonResponse({"ok": False, "msg": "Encoding inválido (UTF-8)"}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "msg": "JSON inválido"}, status=400)
    except Exception:
        return JsonResponse({"ok": False, "msg": "Error leyendo el request"}, status=400)

    if not raw:
        return JsonResponse({"ok": False, "msg": "No llegó ningún código/DNI"}, status=400)

    dni = _extract_dni(raw)
    if not (dni.isdigit() and len(dni) == 8):
        return JsonResponse({"ok": False, "msg": "DNI inválido (debe ser 8 dígitos)"}, status=400)

    try:
        profesor = Profesor.objects.get(dni=dni)
    except Profesor.DoesNotExist:
        return JsonResponse({"ok": False, "msg": "Profesor no encontrado", "dni": dni}, status=404)

    hoy = timezone.localdate()

    if _es_dia_especial(hoy):
        return JsonResponse(
            {
                "ok": False,
                "msg": "Hoy es un día especial institucional. No se registra asistencia.",
            },
            status=400,
        )

    now = timezone.now()

    payload_prof = {
        "dni": profesor.dni,
        "codigo": profesor.codigo,
        "apellidos": profesor.apellidos,
        "nombres": profesor.nombres,
        "condicion": profesor.condicion,
    }

    ip = _get_client_ip(request)
    ua = (request.META.get("HTTP_USER_AGENT") or "")[:255]

    with transaction.atomic():
        qs = (
            Asistencia.objects.select_for_update()
            .filter(profesor=profesor, fecha=hoy, tipo="E")
            .order_by("fecha_hora")
        )

        if qs.exists():
            logger.info(
                "DUPLICADO asistencia dni=%s hoy=%s user=%s ip=%s",
                dni,
                hoy,
                request.user.username,
                ip,
            )
            return JsonResponse(
                {
                    "ok": True,
                    "duplicado": True,
                    "accion": "ninguna",
                    "msg": f"⚠️ Ya registró asistencia hoy: {profesor.apellidos} {profesor.nombres}",
                    "profesor": payload_prof,
                },
                status=200,
            )

        asistencia = Asistencia.objects.create(
            profesor=profesor,
            fecha=hoy,
            fecha_hora=now,
            tipo="E",
            registrado_por=request.user,
            ip=ip,
            user_agent=ua,
        )

    logger.info("OK asistencia dni=%s hoy=%s user=%s ip=%s", dni, hoy, request.user.username, ip)

    return JsonResponse(
        {
            "ok": True,
            "duplicado": False,
            "accion": "asistencia",
            "msg": f"✅ ASISTIÓ registrado correctamente: {profesor.apellidos} {profesor.nombres}",
            "profesor": payload_prof,
            "asistencia": {
                "id": asistencia.id,
                "tipo": asistencia.tipo,
                "fecha": str(asistencia.fecha),
                "fecha_hora": asistencia.fecha_hora.isoformat(),
            },
        },
        status=201,
    )


# =========================================================
# CRON PRIVADO
# =========================================================
@csrf_exempt
@require_GET
def trigger_reporte_asistencia(request):
    token = (request.GET.get("token") or "").strip()
    secret = (getattr(settings, "REPORT_TRIGGER_TOKEN", "") or "").strip()

    if not secret or token != secret:
        return HttpResponseForbidden("Forbidden")

    out = StringIO()
    call_command("enviar_reporte_asistencia", stdout=out, stderr=out)
    texto = out.getvalue()

    return JsonResponse({"ok": True, "msg": "Reporte ejecutado", "output": texto})


# =========================================================
# REGISTRO MANUAL
# =========================================================
@user_passes_test(_in_any_group("HISTORIAL", "JUSTIFICACIONES"), login_url="login")
def registro_manual(request):
    origen = (
        request.GET.get("from")
        or request.POST.get("from")
        or request.session.get("historial_origen")
        or "historial"
    ).strip().lower()

    if origen not in ("historial", "justificaciones"):
        origen = "historial"

    request.session["historial_origen"] = origen

    def _url_volver_historial():
        return f"{reverse('historial_asistencias')}?from={origen}"

    def _volver_historial():
        return redirect(_url_volver_historial())

    def _render(ctx_extra=None):
        ctx = {
            "origen_modulo": origen,
            "url_volver_historial": _url_volver_historial(),
        }
        if ctx_extra:
            ctx.update(ctx_extra)
        return render(request, "asistencias/registro_manual.html", ctx)

    if request.method == "GET":
        return _render()

    accion = (request.POST.get("accion") or "").strip().lower()

    if accion == "buscar":
        dni = (request.POST.get("dni") or "").strip()

        if not (dni.isdigit() and len(dni) == 8):
            messages.error(request, "DNI inválido (debe tener 8 dígitos).")
            return _render()

        profesor = Profesor.objects.filter(dni=dni).first()
        if not profesor:
            messages.error(request, "No se encontró un docente con ese DNI.")
            return _render()

        ahora_local = timezone.localtime(timezone.now())
        return _render(
            {
                "profesor": profesor,
                "fecha_hora_str": ahora_local.strftime("%d/%m/%Y %H:%M:%S"),
            }
        )

    if accion == "aceptar":
        dni = (request.POST.get("dni") or "").strip()

        if not (dni.isdigit() and len(dni) == 8):
            messages.error(request, "DNI inválido.")
            return _volver_historial()

        profesor = Profesor.objects.filter(dni=dni).first()
        if not profesor:
            messages.error(request, "El docente no existe o no fue encontrado.")
            return _volver_historial()

        fecha = timezone.localdate()

        if _es_dia_especial(fecha):
            messages.warning(request, "Hoy es un día especial institucional. No se registra asistencia.")
            return _volver_historial()

        ahora = timezone.now()

        if Asistencia.objects.filter(profesor=profesor, fecha=fecha, tipo="E").exists():
            messages.warning(request, "Ese docente ya tiene asistencia registrada hoy.")
            return _volver_historial()

        try:
            Asistencia.objects.create(
                profesor=profesor,
                fecha=fecha,
                fecha_hora=ahora,
                tipo="E",
                registrado_por=request.user,
                ip=_get_client_ip(request),
                user_agent=(request.META.get("HTTP_USER_AGENT", "") or "")[:255],
            )
        except IntegrityError:
            messages.warning(request, "Ese docente ya tiene asistencia registrada hoy.")
            return _volver_historial()

        messages.success(request, "✅ Asistencia registrada correctamente.")
        return _volver_historial()

    messages.error(request, "Acción no válida.")
    return _volver_historial()


# =========================================================
# PANEL JUSTIFICACIONES
# =========================================================
@user_passes_test(_in_group("JUSTIFICACIONES"), login_url="login")
@require_GET
def panel_justificaciones(request):
    hoy = timezone.localdate()

    fecha_str = (request.GET.get("fecha") or "").strip()
    q = (request.GET.get("q") or "").strip()

    if not fecha_str:
        fecha = hoy - timedelta(days=1)
    else:
        fecha = parse_date(fecha_str) or hoy

    request.session["just_fecha"] = fecha.strftime("%Y-%m-%d")

    dia_especial = (
        DiaEspecial.objects
        .filter(fecha=fecha, activo=True)
        .first()
    )

    profesores_qs = (
        Profesor.objects.only("id", "dni", "codigo", "apellidos", "nombres", "condicion")
        .order_by("apellidos", "nombres")
    )

    if q:
        profesores_qs = profesores_qs.filter(
            Q(dni__icontains=q)
            | Q(codigo__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(nombres__icontains=q)
        )

    profesores = list(profesores_qs)

    asist_ids = set(
        Asistencia.objects.filter(fecha=fecha, tipo="E").values_list("profesor_id", flat=True)
    )

    just_qs = (
        JustificacionAsistencia.objects.only(
            "id",
            "profesor_id",
            "fecha",
            "tipo",
            "detalle",
            "archivo",
        )
        .filter(fecha=fecha)
    )

    just_map = {j.profesor_id: j for j in just_qs}

    rows = []
    c_asistio = 0
    c_just = 0
    c_falto = 0
    c_especial = 0

    for profesor in profesores:
        if profesor.id in asist_ids:
            estado_key = "ASISTIO"
            estado = "ASISTIÓ"
            estado_detalle = "Asistencia registrada"
            justificacion_info = None
            c_asistio += 1
        else:
            j = just_map.get(profesor.id)
            if j:
                tipo_label = j.get_tipo_display() if hasattr(j, "get_tipo_display") else (j.tipo or "")
                estado_key = "JUSTIFICADO"
                estado = f"JUSTIFICADO ({tipo_label})"
                estado_detalle = (j.detalle or "").strip()
                justificacion_info = {
                    "id": j.id,
                    "tipo": j.tipo,
                    "tipo_label": tipo_label,
                    "detalle": j.detalle or "",
                    "archivo_url": _safe_file_url(j.archivo),
                }
                c_just += 1
            elif dia_especial:
                tipo_display = _tipo_display_dia_especial(dia_especial)
                descripcion = (dia_especial.descripcion or "").strip()

                estado_key = "DIA_ESPECIAL"
                estado = tipo_display.upper()
                estado_detalle = descripcion or "Día especial institucional"
                justificacion_info = None
                c_especial += 1
            else:
                estado_key = "FALTO"
                estado = "FALTÓ"
                estado_detalle = "Sin asistencia ni justificación"
                justificacion_info = None
                c_falto += 1

        rows.append(
            {
                "profesor": profesor,
                "estado": estado,
                "estado_key": estado_key,
                "estado_detalle": estado_detalle,
                "justificacion": justificacion_info,
                "es_dia_especial": bool(dia_especial),
            }
        )

    can_historial = (
        request.user.is_superuser
        or request.user.groups.filter(name="HISTORIAL").exists()
        or request.user.groups.filter(name="JUSTIFICACIONES").exists()
    )

    return render(
        request,
        "asistencias/justificaciones.html",
        {
            "fecha": fecha,
            "q": q,
            "rows": rows,
            "can_historial": can_historial,
            "dia_especial": dia_especial,
            "resumen": {
                "asistio": c_asistio,
                "justificado": c_just,
                "falto": c_falto,
                "especial": c_especial,
                "total": c_asistio + c_just + c_falto + c_especial,
            },
        },
    )


# =========================================================
# SET JUSTIFICACIÓN
# =========================================================
@require_POST
@user_passes_test(_in_group("JUSTIFICACIONES"), login_url="login")
def set_justificacion(request):
    accion = (request.POST.get("accion") or "").strip().lower()
    profesor_id = (request.POST.get("profesor_id") or "").strip()
    fecha_str = (request.POST.get("fecha") or "").strip()
    tipo = (request.POST.get("tipo") or "DM").strip().upper()
    detalle = (request.POST.get("detalle") or "").strip()
    archivo = request.FILES.get("archivo")

    redirect_url = (
        f"/asistencia/justificaciones/?fecha={fecha_str}"
        if fecha_str else "/asistencia/justificaciones/"
    )

    if accion != "set":
        messages.error(request, "Acción inválida.")
        return redirect(redirect_url)

    fecha = parse_date(fecha_str)
    if not fecha:
        messages.error(request, "Fecha inválida.")
        return redirect(redirect_url)

    if _es_dia_especial(fecha):
        messages.warning(request, "Ese día está marcado como día especial. No se requiere justificación.")
        return redirect(redirect_url)

    try:
        profesor = Profesor.objects.get(id=profesor_id)
    except Profesor.DoesNotExist:
        messages.error(request, "Profesor no encontrado.")
        return redirect(redirect_url)

    tipo_ok = tipo if tipo in ("DM", "C", "P", "O") else "DM"
    ip = _get_client_ip(request)
    ua = (request.META.get("HTTP_USER_AGENT") or "")[:255]

    if Asistencia.objects.filter(profesor=profesor, fecha=fecha, tipo="E").exists():
        messages.warning(request, "🛑 Ya tiene ASISTENCIA ese día. No se registró justificación.")
        return redirect(redirect_url)

    if JustificacionAsistencia.objects.filter(profesor=profesor, fecha=fecha).exists():
        messages.warning(
            request,
            "✅ Este docente ya fue justificado en esta fecha. (Solo se puede editar en el Admin).",
        )
        return redirect(redirect_url)

    if archivo:
        nombre_original = (archivo.name or "").strip()
        nombre_lower = nombre_original.lower()
        ctype = (getattr(archivo, "content_type", "") or "").lower()
        size = getattr(archivo, "size", 0) or 0

        if not nombre_lower.endswith(".pdf"):
            messages.error(request, "El archivo debe terminar en .pdf")
            return redirect(redirect_url)

        if ctype and ctype != "application/pdf":
            messages.error(request, f"El archivo debe ser PDF (content_type recibido: {ctype}).")
            return redirect(redirect_url)

        if size > 10 * 1024 * 1024:
            messages.error(request, "El PDF es muy pesado (máx. 10 MB).")
            return redirect(redirect_url)

    try:
        with transaction.atomic():
            just_kwargs = {
                "profesor": profesor,
                "fecha": fecha,
                "tipo": tipo_ok,
                "detalle": detalle,
                "creado_por": request.user,
                "actualizado_por": request.user,
            }

            if archivo:
                cloudinary_error = None

                if CLOUDINARY_AVAILABLE:
                    try:
                        folder = f"justificaciones/{fecha.year}/{fecha.month:02d}"
                        cloudinary.uploader.upload(
                            archivo,
                            resource_type="raw",
                            folder=folder,
                            use_filename=True,
                            unique_filename=True,
                            overwrite=False,
                        )
                        try:
                            archivo.seek(0)
                        except Exception:
                            pass
                    except Exception as ex:
                        cloudinary_error = str(ex)[:220]
                        try:
                            archivo.seek(0)
                        except Exception:
                            pass

                just_kwargs["archivo"] = archivo

                if cloudinary_error:
                    messages.warning(
                        request,
                        f"⚠️ Cloudinary falló; se guardó el PDF en el almacenamiento del sistema. ({cloudinary_error})",
                    )

            JustificacionAsistencia.objects.create(**just_kwargs)

            Asistencia.objects.update_or_create(
                profesor=profesor,
                fecha=fecha,
                tipo="J",
                defaults={
                    "fecha_hora": timezone.now(),
                    "motivo": tipo_ok,
                    "detalle": detalle,
                    "registrado_por": request.user,
                    "ip": ip,
                    "user_agent": ua,
                },
            )

        messages.success(request, "✅ Justificación guardada correctamente.")
        return redirect(redirect_url)

    except IntegrityError:
        messages.warning(request, "✅ Ya existía una justificación para ese docente en esa fecha.")
        return redirect(redirect_url)

    except Exception as e:
        messages.error(
            request,
            f"Error guardando justificación/PDF: {type(e).__name__} - {str(e)[:250]}",
        )
        return redirect(redirect_url)


# =========================================================
# ESTADÍSTICAS PRIVADAS
# =========================================================
def _build_private_stats(fecha_inicio, fecha_fin, q="", condicion=""):
    profesores_qs = Profesor.objects.all().order_by("apellidos", "nombres")

    if q:
        profesores_qs = profesores_qs.filter(
            Q(dni__icontains=q)
            | Q(codigo__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(nombres__icontains=q)
        )

    if condicion in ("N", "C"):
        profesores_qs = profesores_qs.filter(condicion__iexact=condicion)

    profesores = list(profesores_qs)
    profesor_ids = [p.id for p in profesores]

    dias_especiales = _dias_especiales_dict(fecha_inicio, fecha_fin)

    dias_habiles = []
    cur = fecha_inicio
    while cur <= fecha_fin:
        if cur.weekday() < 5 and cur not in dias_especiales:
            dias_habiles.append(cur)
        cur += timedelta(days=1)

    entradas_set = set(
        Asistencia.objects.filter(
            profesor_id__in=profesor_ids,
            fecha__range=(fecha_inicio, fecha_fin),
            tipo="E",
        ).values_list("profesor_id", "fecha")
    )

    just_set = set(
        JustificacionAsistencia.objects.filter(
            profesor_id__in=profesor_ids,
            fecha__range=(fecha_inicio, fecha_fin),
        ).values_list("profesor_id", "fecha")
    )

    just_asist_set = set(
        Asistencia.objects.filter(
            profesor_id__in=profesor_ids,
            fecha__range=(fecha_inicio, fecha_fin),
            tipo="J",
        ).values_list("profesor_id", "fecha")
    )

    rows = []
    total_asistio = 0
    total_justifico = 0
    total_falto = 0

    for profesor in profesores:
        asistio = 0
        justifico = 0
        falto = 0

        for dia in dias_habiles:
            key = (profesor.id, dia)

            if key in entradas_set:
                asistio += 1
            elif key in just_set or key in just_asist_set:
                justifico += 1
            else:
                falto += 1

        total_dias = len(dias_habiles)
        porcentaje = round((asistio / total_dias) * 100, 2) if total_dias else 0

        rows.append(
            {
                "profesor": profesor,
                "asistio": asistio,
                "justifico": justifico,
                "falto": falto,
                "total_dias": total_dias,
                "porcentaje": porcentaje,
            }
        )

        total_asistio += asistio
        total_justifico += justifico
        total_falto += falto

    docentes_total = len(rows)
    base_total = total_asistio + total_justifico + total_falto
    porcentaje_general = round((total_asistio / base_total) * 100, 2) if base_total else 0

    return {
        "rows": rows,
        "dias_habiles": dias_habiles,
        "dias_especiales": dias_especiales,
        "docentes_total": docentes_total,
        "total_asistio": total_asistio,
        "total_justifico": total_justifico,
        "total_falto": total_falto,
        "porcentaje_general": porcentaje_general,
    }


@login_required
def estadisticas_privadas(request):
    if not _is_private_owner(request.user):
        return HttpResponseForbidden("No tienes permiso para acceder a esta sección.")

    hoy = timezone.localdate()

    fecha_inicio = parse_date((request.GET.get("inicio") or "").strip())
    fecha_fin = parse_date((request.GET.get("fin") or "").strip())
    q = (request.GET.get("q") or "").strip()
    condicion = (request.GET.get("condicion") or "").strip().upper()

    if not fecha_fin:
        fecha_fin = hoy

    if not fecha_inicio:
        fecha_inicio = fecha_fin - timedelta(days=6)

    if fecha_inicio > fecha_fin:
        fecha_inicio, fecha_fin = fecha_fin, fecha_inicio

    stats = _build_private_stats(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        q=q,
        condicion=condicion,
    )

    return render(
        request,
        "asistencias/estadisticas_privadas.html",
        {
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "q": q,
            "condicion": condicion,
            **stats,
        },
    )


@login_required
def exportar_estadisticas_privadas_excel(request):
    if not _is_private_owner(request.user):
        return HttpResponseForbidden("No tienes permiso para exportar esta información.")

    hoy = timezone.localdate()

    fecha_inicio = parse_date((request.GET.get("inicio") or "").strip())
    fecha_fin = parse_date((request.GET.get("fin") or "").strip())
    q = (request.GET.get("q") or "").strip()
    condicion = (request.GET.get("condicion") or "").strip().upper()

    if not fecha_fin:
        fecha_fin = hoy

    if not fecha_inicio:
        fecha_inicio = fecha_fin - timedelta(days=6)

    if fecha_inicio > fecha_fin:
        fecha_inicio, fecha_fin = fecha_fin, fecha_inicio

    stats = _build_private_stats(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        q=q,
        condicion=condicion,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas privadas"

    navy = "7F1D1D"
    red = "B91C1C"
    green_bg = "DCFCE7"
    amber_bg = "FEF3C7"
    red_bg = "FEE2E2"
    white = "FFFFFF"

    thin = Side(style="thin", color="CBD5E1")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    ws["A1"] = "PROYECTO MANHATTAN - ESTADÍSTICAS PRIVADAS DE ASISTENCIA"
    ws["A1"].font = Font(bold=True, size=15, color=navy)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Rango: {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')} | Lunes a viernes | Excluye días especiales"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Docente", "Condición", "Asistió", "Justificó", "Faltó", "Total días", "% Asistencia"]
    row_header = 4

    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_header, column=i, value=header)
        cell.fill = PatternFill("solid", fgColor=red)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all

    row = row_header + 1

    for item in stats["rows"]:
        profesor = item["profesor"]
        docente = f"{(profesor.apellidos or '').strip()}, {(profesor.nombres or '').strip()}".strip().strip(",")

        values = [
            docente,
            (profesor.condicion or "").upper(),
            item["asistio"],
            item["justifico"],
            item["falto"],
            item["total_dias"],
            item["porcentaje"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border_all
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col == 3:
                cell.fill = PatternFill("solid", fgColor=green_bg)
            elif col == 4:
                cell.fill = PatternFill("solid", fgColor=amber_bg)
            elif col == 5:
                cell.fill = PatternFill("solid", fgColor=red_bg)
            elif col == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        row += 1

    ws["A3"] = f"Docentes evaluados: {stats['docentes_total']}"
    ws["D3"] = f"Asistió: {stats['total_asistio']}"
    ws["E3"] = f"Justificó: {stats['total_justifico']}"
    ws["F3"] = f"Faltó: {stats['total_falto']}"
    ws["G3"] = f"% General: {stats['porcentaje_general']}%"

    widths = [38, 14, 12, 12, 12, 12, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    filename = f"estadisticas_privadas_{fecha_inicio.strftime('%Y-%m-%d')}_a_{fecha_fin.strftime('%Y-%m-%d')}.xlsx"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response